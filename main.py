from openpyxl import load_workbook
from openpyxl.styles import Font
import argparse

parser = argparse.ArgumentParser(
    description='Import notes column from xlsx to another xlsx.')

parser.add_argument('file1', type=str,
                    help='the previous file to extract notes')
parser.add_argument('file2', type=str,
                    help='the file to append previous notes')
parser.add_argument('start', type=str,
                    help='the name of the first header column (Last Name)')


def create_key(row):
    lastName = row[0].value if row[0].value else ''
    firstName = row[1].value if row[1].value else ''
    org = row[2].value if row[2].value else ''
    return '_'.join([lastName, firstName, org])


def get_data(file_name, start_name):
    """ Gets all data starting from the row of the start name.
        Start name is the value in a cell.
    """
    wb = load_workbook(file_name)
    sheet = wb.worksheets[0]
    start = find_start(sheet, start_name)
    # Add better logic to find notes. We assume notes are the last and second to last in the xlsx
    notes_column = sheet.max_column - 1
    data = {}
    for start in range(start, sheet.max_row):
        key = create_key(list(sheet.rows)[start])
        row = list(sheet.rows)[start]
        data[key] = (row[notes_column].value,
                     row[notes_column - 1].value)
    return data


def find_start(sheet, start_name):
    """  Return index of the row with the start name """
    for i in range(1, sheet.max_row):
        if sheet.cell(row=i, column=1).value == start_name:
            return i
    raise ValueError('Start name not found in file1!')


def insert_previous_notes(sheet, start, notes_column):
    rows = list(sheet.rows)
    for start in range(start, sheet.max_row+1):
        key = create_key(rows[start-1])
        print(start, sheet.max_row)
        val = key in data
        if val is True:
            curr_notes_column = notes_column
            for i in range(0, len(data[key])):
                sheet.cell(row=start, column=curr_notes_column).value = str(
                    data[key][i])
                curr_notes_column += 1
        else:
            for cell in sheet[start]:
                cell.font = Font(color='00FF0000', italic=True)
    return sheet


def process_data(data, file_path, start_name):
    wb = load_workbook(file_path)
    sheet = wb.worksheets[0]
    start = find_start(sheet, start_name)
    # Add better logic to find notes. We assume notes are the last and second to last in the xlsx 
    notes_column = sheet.max_column + 1
    sheet.cell(row=start, column=notes_column, value='Notes_imported').font \
        = Font(color='00000000', italic=True, bold=True)
    sheet = insert_previous_notes(sheet, start+1, notes_column)
    wb.save(file_path.split('.')[0] + '_formated.xlsx')


if __name__ == '__main__':
    args = parser.parse_args()

    print('Getting data from {}'.format(args.file1))
    data = get_data(args.file1, args.start)

    print('Formating {} with previous notes'.format(args.file2))
    process_data(data, args.file2, args.start)
    print('Job Complete')
